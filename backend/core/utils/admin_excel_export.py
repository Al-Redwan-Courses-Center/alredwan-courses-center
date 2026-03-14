#!/usr/bin/env python3
"""
Excel Export Mixin for Django Admin

This mixin adds Excel export functionality to any Django admin class.
It exports the exact data that appears in the admin list view, including
all applied filters, search queries, and ordering.

Usage:
    from core.utils.admin_excel_export import ExcelExportMixin
    
    @admin.register(MyModel)
    class MyModelAdmin(ExcelExportMixin, admin.ModelAdmin):
        list_display = ['field1', 'field2', ...]
        excel_export_fields = ['field1', 'field2', ...]  # Optional
        excel_export_exclude = ['password']  # Optional
"""

from django.contrib import admin
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, time
import re


class ExcelExportMixin:
    """
    Mixin to add Excel export functionality to Django admin.

    Attributes:
        excel_export_fields (list): Fields to export. If not specified, uses list_display.
        excel_export_exclude (list): Fields to exclude from export.
        excel_filename (str): Custom filename for the export (without extension).
    """

    excel_export_fields = None
    excel_export_exclude = []
    excel_filename = None
    
    # Internal Django admin fields that should never be exported (but can appear in list_display)
    _internal_fields = ['action_checkbox', '__str__', '__repr__']

    def get_actions(self, request):
        """Add export to Excel action to the admin actions."""
        actions = super().get_actions(request)
        if actions is None:
            actions = {}
        actions['export_to_excel'] = (
            self.export_to_excel,
            'export_to_excel',
            _('تصدير إلى Excel (Export to Excel)')
        )
        return actions

    def get_list_display_links(self, request, list_display):
        """Ensure at least one non-checkbox column links to the change page."""
        links = super().get_list_display_links(request, list_display)
        if not list_display:
            return links

        if links and any(field != 'action_checkbox' for field in links):
            return links

        for field in list_display:
            if field != 'action_checkbox':
                return (field,)

        return links

    def changelist_view(self, request, extra_context=None):
        """Add export button to the changelist view and handle direct export."""
        extra_context = extra_context or {}
        extra_context['has_excel_export'] = True

        # Check if this is a direct export request (via URL parameter)
        if request.GET.get('export') == 'excel':
            # Create a mutable copy of GET parameters and remove 'export'
            mutable_get = request.GET.copy()
            mutable_get.pop('export', None)

            # Create a modified request object with cleaned GET parameters
            # We need to temporarily replace request.GET
            original_get = request.GET
            request.GET = mutable_get

            try:
                result = self.direct_export_to_excel(request)
            finally:
                # Restore original GET parameters
                request.GET = original_get

            return result

        return super().changelist_view(request, extra_context)

    def clean_sheet_name(self, name):
        r"""
        Clean sheet name by removing invalid characters for Excel.
        Excel sheet names cannot contain: / \ ? * [ ] :
        and must be 31 characters or less.
        """
        # Remove invalid characters
        invalid_chars = r'[/\\?*\[\]:]'
        clean_name = re.sub(invalid_chars, '_', str(name))

        # Trim to 31 characters (Excel limit)
        clean_name = clean_name[:31]

        # Remove leading/trailing spaces and underscores
        clean_name = clean_name.strip(' _')

        # If empty after cleaning, use default name
        if not clean_name:
            clean_name = 'Export'

        return clean_name

    def direct_export_to_excel(self, request):
        """
        Direct export handler that doesn't require item selection.
        Called when ?export=excel is in the URL.
        """
        # Get the changelist to respect filters and ordering
        ChangeList = self.get_changelist_instance(request)

        # Use the filtered queryset from changelist
        filtered_queryset = ChangeList.get_queryset(request)

        # Check if there's any data to export
        if not filtered_queryset.exists():
            from django.contrib import messages
            messages.warning(request, _(
                'لا توجد بيانات للتصدير (No data to export)'))
            # Redirect back to changelist
            from django.shortcuts import redirect
            return redirect(request.path)

        # Create workbook
        wb = Workbook()
        ws = wb.active

        # Set worksheet title with cleaned name
        model_name = self.model._meta.verbose_name_plural
        ws.title = self.clean_sheet_name(model_name)

        # Get fields to export
        fields = self.get_export_fields()

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_num, field_info in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_num)
            # Convert label to string to avoid openpyxl conversion errors
            cell.value = str(field_info['label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        for row_num, obj in enumerate(filtered_queryset, 2):
            for col_num, field_info in enumerate(fields, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = self.get_field_value(obj, field_info)
                # Ensure value is properly converted to string if needed
                if value is None:
                    cell.value = ''
                elif isinstance(value, (str, int, float, bool)):
                    cell.value = value
                else:
                    cell.value = str(value)
                cell.border = border

                # Right-to-left alignment for Arabic text
                if isinstance(value, str) and self.contains_arabic(value):
                    cell.alignment = Alignment(
                        horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center")

        # Auto-adjust column widths
        for col_num, field_info in enumerate(fields, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(field_info['label']))

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                    min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze first row
        ws.freeze_panes = 'A2'

        # Prepare response
        filename = self.get_export_filename()
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        full_filename = f"{filename}_{timestamp}.xlsx"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{full_filename}"'

        wb.save(response)
        return response

    def export_to_excel(self, request, queryset):
        """
        Export the filtered/ordered queryset to Excel.
        This respects all filters, search, and ordering applied in the admin.
        Note: This action ignores the selected items and exports ALL filtered results.
        """
        # Get the changelist to respect filters and ordering
        # We use changelist instead of the queryset parameter to get ALL filtered items
        ChangeList = self.get_changelist_instance(request)

        # Use the filtered queryset from changelist (ignores selection)
        filtered_queryset = ChangeList.get_queryset(request)

        # Check if there's any data to export
        if not filtered_queryset.exists():
            from django.contrib import messages
            messages.warning(request, _(
                'لا توجد بيانات للتصدير (No data to export)'))
            return None

        # Create workbook
        wb = Workbook()
        ws = wb.active

        # Set worksheet title with cleaned name
        model_name = self.model._meta.verbose_name_plural
        ws.title = self.clean_sheet_name(model_name)

        # Get fields to export
        fields = self.get_export_fields()

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_num, field_info in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_num)
            # Convert label to string to avoid openpyxl conversion errors
            cell.value = str(field_info['label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        row_count = 0
        for row_num, obj in enumerate(filtered_queryset, 2):
            row_count += 1
            for col_num, field_info in enumerate(fields, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = self.get_field_value(obj, field_info)
                # Ensure value is properly converted to string if needed
                if value is None:
                    cell.value = ''
                elif isinstance(value, (str, int, float, bool)):
                    cell.value = value
                else:
                    cell.value = str(value)
                cell.border = border

                # Right-to-left alignment for Arabic text
                if isinstance(value, str) and self.contains_arabic(value):
                    cell.alignment = Alignment(
                        horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center")

        # Auto-adjust column widths
        for col_num, field_info in enumerate(fields, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(field_info['label']))

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                    min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze first row
        ws.freeze_panes = 'A2'

        # Prepare response
        filename = self.get_export_filename()
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        full_filename = f"{filename}_{timestamp}.xlsx"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{full_filename}"'

        wb.save(response)

        # Show success message
        from django.contrib import messages
        messages.success(
            request,
            _('تم تصدير {} صف بنجاح (Successfully exported {} rows)').format(
                row_count, row_count)
        )

        return response

    # Mark the action to work without selecting items
    export_to_excel.short_description = _('تصدير إلى Excel (Export to Excel)')
    export_to_excel.allowed_permissions = ()  # No special permissions needed

    def get_changelist_instance(self, request):
        """Get the ChangeList instance to access filtered queryset."""
        from django.contrib.admin.views.main import ChangeList

        list_display = self.get_list_display(request)
        list_display_links = self.get_list_display_links(request, list_display)
        list_filter = self.get_list_filter(request)
        search_fields = self.get_search_fields(request)
        list_select_related = self.get_list_select_related(request)

        # Build ChangeList arguments - handle optional parameters for compatibility
        changelist_kwargs = {
            'request': request,
            'model': self.model,
            'list_display': list_display,
            'list_display_links': list_display_links,
            'list_filter': list_filter,
            'date_hierarchy': self.date_hierarchy,
            'search_fields': search_fields,
            'list_select_related': list_select_related,
            'list_per_page': self.list_per_page,
            'list_max_show_all': self.list_max_show_all,
            'list_editable': self.list_editable,
            'model_admin': self,
            'sortable_by': self.get_sortable_by(request),
        }

        # Add search_help_text - provide default if method doesn't exist
        if hasattr(self, 'get_search_help_text'):
            changelist_kwargs['search_help_text'] = self.get_search_help_text(
                request)
        else:
            # Provide empty string as default for Django 5.2+
            changelist_kwargs['search_help_text'] = ''

        return ChangeList(**changelist_kwargs)

    def get_export_fields(self):
        """
        Get the list of fields to export with their labels.
        Returns a list of dicts with 'name', 'label', and 'is_method' keys.
        """
        if self.excel_export_fields:
            field_names = self.excel_export_fields
        else:
            # Filter out internal Django admin fields and excluded fields
            field_names = [
                f for f in self.list_display 
                if f not in self.excel_export_exclude
            ]

        fields = []
        for field_name in field_names:
            # Skip internal fields and excluded fields
            if field_name in self.excel_export_exclude or field_name in self._internal_fields:
                continue

            field_info = {
                'name': field_name,
                'label': self.get_field_label(field_name),
                'is_method': hasattr(self, field_name) or hasattr(self.model, field_name)
            }
            fields.append(field_info)

        return fields

    def get_field_label(self, field_name):
        """Get the human-readable label for a field."""
        # Check if it's a method on the admin class
        if hasattr(self, field_name):
            method = getattr(self, field_name)
            if hasattr(method, 'short_description'):
                return method.short_description
            if hasattr(method, 'admin_order_field'):
                return field_name.replace('_', ' ').title()

        # Check if it's a model field
        try:
            field = self.model._meta.get_field(field_name)
            return field.verbose_name.title()
        except:
            pass

        # Check if it's a model method with admin decorator
        if hasattr(self.model, field_name):
            method = getattr(self.model, field_name)
            if hasattr(method, 'short_description'):
                return method.short_description

        # Fallback to formatted field name
        return field_name.replace('_', ' ').title()

    def get_field_value(self, obj, field_info):
        """
        Get the value of a field for an object.
        Handles model fields, methods, and admin display methods.
        """
        field_name = field_info['name']

        # Try admin method first
        if hasattr(self, field_name):
            method = getattr(self, field_name)
            if callable(method):
                try:
                    value = method(obj)
                    return self.clean_html(value)
                except:
                    return ''

        # Try model method
        if hasattr(obj, field_name):
            attr = getattr(obj, field_name)
            if callable(attr):
                try:
                    value = attr()
                    return self.format_value(value)
                except:
                    return ''
            else:
                return self.format_value(attr)

        return ''

    def format_value(self, value):
        """Format a value for Excel export."""
        if value is None:
            return ''
        if isinstance(value, bool):
            return 'نعم' if value else 'لا'
        if isinstance(value, (datetime, date)):
            return value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
        if isinstance(value, time):
            return value.strftime('%H:%M:%S')
        return str(value)

    def clean_html(self, value):
        """Remove HTML tags from a value."""
        if not isinstance(value, str):
            return self.format_value(value)

        import re
        # Remove HTML tags
        clean_text = re.sub(r'<[^>]+>', '', value)
        # Decode HTML entities
        import html
        clean_text = html.unescape(clean_text)
        return clean_text.strip()

    def contains_arabic(self, text):
        """Check if text contains Arabic characters."""
        if not isinstance(text, str):
            return False
        arabic_pattern = r'[\u0600-\u06FF]'
        import re
        return bool(re.search(arabic_pattern, text))

    def get_export_filename(self):
        """Get the filename for the export."""
        if self.excel_filename:
            return self.excel_filename
        return self.model._meta.verbose_name_plural.replace(' ', '_')
