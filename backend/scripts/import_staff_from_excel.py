#!/usr/bin/env python
"""
Script to import Instructors and Supervisors from Excel file (Arabic headers support)
This script can be run while the Django server is running.

Features:
- Supports Arabic column headers from Google Forms
- Auto-generates fingerprint IDs
- Auto-generates secure random passwords
- Downloads ID images from Google Drive and uploads to Cloudinary
- Skips duplicates

Usage:
    python scripts/import_staff_from_excel.py path/to/excel_file.xlsx
"""

import os
import sys
import django
from pathlib import Path
from datetime import datetime
from decimal import Decimal
import re
import requests
from io import BytesIO
import secrets
import string

# Setup Django environment
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'Redwan_courses_center.settings')
django.setup()

from django.db import transaction
from django.core.exceptions import ValidationError
from django.contrib.auth.hashers import make_password
from django.core.files.uploadedfile import InMemoryUploadedFile
from users.models import CustomUser, Instructor
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import cloudinary.uploader


class StaffImporter:
    """Import staff (instructors and supervisors) from Excel files with Arabic headers"""
    
    # Column mapping: Arabic header -> English field name
    COLUMN_MAPPING = {
        'الاسم الأول': 'first_name_1',
        'الاسم الثاني': 'first_name_2',
        'الاسم الثالث': 'last_name_1',
        'اسم الرابع': 'last_name_2',
        'البريد الإلكتروني': 'email',
        'تاريخ الميلاد': 'dob',
        'الجنس': 'gender',
        'رقم الهاتف الأساسي (رقم 1)': 'phone_number1',
        'رقم هاتف إضافي (رقم 2) - اختياري': 'phone_number2',
        'رقم هاتف إضافي (رقم 2) - اختياري 2': 'phone_number2_alt',
        'نوع الهوية': 'identity_type',
        'رقم الهوية / الوثيقة': 'identity_number',
        'تحميل صورة الواجهة الأمامية لوثيقة الهوية': 'nid_front_url',
        'تحميل صورة الواجهة الخلفية لوثيقة الهوية': 'nid_back_url',
        'العنوان بالتفصيل': 'address',
        'الموقع الجغرافي (المدينة/المنطقة)': 'location',
        'الدور في الأكاديمية': 'role',
        'الراتب الشهري المتوقع/المتفق عليه': 'monthly_salary',
        'نبذة شخصية وسيرة ذاتية مختصرة (Bio)': 'bio',
        'تعيين كلمة مرور مؤقتة للنظام الداخلي': 'password',
    }
    
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.stats = {
            'total': 0,
            'success': 0,
            'skipped': 0,
            'errors': 0
        }
        self.error_messages = []
        self.fingerprint_counter = self.get_next_fingerprint_number()
        self.imported_users = []  # Store imported user data for Excel export
    
    def get_next_fingerprint_number(self):
        """Get the next available fingerprint ID number"""
        # Get the highest existing fingerprint ID
        last_instructor = Instructor.objects.filter(
            fingerprint_id__isnull=False
        ).exclude(fingerprint_id='').order_by('-fingerprint_id').first()
        
        if last_instructor and last_instructor.fingerprint_id:
            # Extract number from fingerprint_id (e.g., "FP001" -> 1)
            match = re.search(r'\d+', last_instructor.fingerprint_id)
            if match:
                return int(match.group()) + 1
        
        return 1  # Start from 1 if no existing IDs
    
    def generate_fingerprint_id(self):
        """Generate a unique fingerprint ID"""
        fingerprint_id = f"FP{self.fingerprint_counter:04d}"
        self.fingerprint_counter += 1
        return fingerprint_id
    
    def generate_secure_password(self, length=12):
        """
        Generate a secure random password that meets Django's validation requirements:
        - At least 8 characters (we use 12 for better security)
        - Contains uppercase, lowercase, numbers, and special characters
        - Not similar to user attributes
        - Not a common password
        - Not entirely numeric
        """
        # Define character sets
        uppercase = string.ascii_uppercase
        lowercase = string.ascii_lowercase
        digits = string.digits
        special_chars = '@$!%*?&#'
        
        # Ensure at least one character from each set
        password = [
            secrets.choice(uppercase),
            secrets.choice(uppercase),  # Add extra uppercase
            secrets.choice(lowercase),
            secrets.choice(lowercase),  # Add extra lowercase
            secrets.choice(digits),
            secrets.choice(digits),  # Add extra digits
            secrets.choice(special_chars),
        ]
        
        # Fill the rest with random characters from all sets
        all_chars = uppercase + lowercase + digits + special_chars
        password += [secrets.choice(all_chars) for _ in range(length - len(password))]
        
        # Shuffle to avoid predictable patterns
        secrets.SystemRandom().shuffle(password)
        
        return ''.join(password)
    
    def extract_google_drive_file_id(self, url):
        """Extract file ID from Google Drive URL"""
        if not url:
            return None
        
        # Pattern: https://drive.google.com/open?id=FILE_ID
        # Pattern: https://drive.google.com/file/d/FILE_ID/view
        patterns = [
            r'id=([a-zA-Z0-9_-]+)',
            r'/d/([a-zA-Z0-9_-]+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        
        return None
    
    def download_google_drive_image(self, url):
        """Download image from Google Drive and return as file-like object"""
        if not url or not isinstance(url, str):
            return None
        
        try:
            file_id = self.extract_google_drive_file_id(url)
            if not file_id:
                print(f"  ⚠️  Could not extract file ID from URL: {url[:50]}...")
                return None
            
            # Google Drive direct download URL
            download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
            
            print(f"  📥 Downloading image from Google Drive...")
            response = requests.get(download_url, timeout=30)
            
            if response.status_code == 200:
                # Check if it's actually an image
                content_type = response.headers.get('content-type', '')
                if 'image' not in content_type.lower() and len(response.content) < 1000:
                    print(f"  ⚠️  Downloaded content doesn't appear to be an image")
                    return None
                
                return BytesIO(response.content)
            else:
                print(f"  ⚠️  Failed to download image (Status: {response.status_code})")
                return None
                
        except Exception as e:
            print(f"  ⚠️  Error downloading image: {str(e)}")
            return None
    
    def upload_to_cloudinary(self, image_data, folder, public_id):
        """Upload image to Cloudinary"""
        if not image_data:
            return None
        
        try:
            print(f"  ☁️  Uploading to Cloudinary...")
            result = cloudinary.uploader.upload(
                image_data,
                folder=folder,
                public_id=public_id,
                resource_type='image',
                transformation={
                    'width': 1200,
                    'crop': 'limit',
                    'quality': 'auto:good',
                    'fetch_format': 'auto',
                }
            )
            print(f"  ✅ Uploaded to Cloudinary successfully")
            return result['public_id']
        except Exception as e:
            print(f"  ⚠️  Error uploading to Cloudinary: {str(e)}")
            return None
    
    def normalize_gender(self, value):
        """Normalize gender values (support Arabic and English)"""
        if not value:
            return None
        value = str(value).strip().lower()
        male_values = ['male', 'ذكر', 'm', 'ذكور']
        female_values = ['female', 'أنثى', 'f', 'إناث', 'انثى']
        
        if value in male_values:
            return 'male'
        elif value in female_values:
            return 'female'
        return None
    
    def normalize_role(self, value):
        """Normalize role values (support Arabic and English)"""
        if not value:
            return None
        value = str(value).strip().lower()
        instructor_values = ['instructor', 'مدرس', 'teacher', 'معلم']
        supervisor_values = ['supervisor', 'مشرف']
        
        if value in instructor_values:
            return 'instructor'
        elif value in supervisor_values:
            return 'supervisor'
        return None
    
    def normalize_instructor_type(self, role):
        """Determine instructor type based on role"""
        if role == 'supervisor':
            return 'supervisor'
        return 'normal'
    
    def normalize_identity_type(self, value):
        """Normalize identity type values"""
        if not value:
            return 'nid'
        value = str(value).strip().lower()
        if value in ['nid', 'national', 'بطاقة', 'وطنية', 'بطاقة هوية وطنية', 'بطاقة وطنية']:
            return 'nid'
        elif value in ['passport', 'جواز', 'جواز سفر']:
            return 'passport'
        elif value in ['driving', 'license', 'رخصة', 'رخصة قيادة']:
            return 'other'
        return 'other'
    
    def clean_phone_number(self, value):
        """Clean and format phone number"""
        if not value:
            return None
        
        # Convert to string and remove spaces, dashes, parentheses
        phone = str(value).strip().replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        
        # If it starts with 0, replace with +20 (Egypt)
        if phone.startswith('0'):
            phone = '+20' + phone[1:]
        
        # If it doesn't start with +, add +20
        if not phone.startswith('+'):
            phone = '+20' + phone
        
        return phone
    
    def parse_date(self, value):
        """Parse date from various formats"""
        if isinstance(value, datetime):
            return value.date()
        
        if not value:
            return None
        
        value = str(value).strip()
        
        # Handle datetime string format (e.g., "2026-02-04 00:00:00")
        if ' ' in value:
            value = value.split(' ')[0]  # Take only the date part
        
        # Try different date formats
        formats = [
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        
        raise ValueError(f"Invalid date format: {value}")
    
    def get_cell_value(self, row, column_name):
        """Safely get cell value from row"""
        value = row.get(column_name)
        if value is None or str(value).strip() == '':
            return None
        return str(value).strip()
    
    def check_duplicate(self, phone_number1, email, identity_number, fingerprint_id):
        """Check if user already exists"""
        # Check phone number
        if CustomUser.objects.filter(phone_number1=phone_number1).exists():
            return f"Phone number {phone_number1} already exists"
        
        # Check email
        if email and CustomUser.objects.filter(email=email).exists():
            return f"Email {email} already exists"
        
        # Check identity number
        if identity_number and CustomUser.objects.filter(identity_number=identity_number).exists():
            return f"Identity number {identity_number} already exists"
        
        # Check fingerprint ID
        if fingerprint_id and Instructor.objects.filter(fingerprint_id=fingerprint_id).exists():
            return f"Fingerprint ID {fingerprint_id} already exists"
        
        return None
    
    def import_row(self, row_data, row_number):
        """Import a single row of data"""
        try:
            # Extract required fields
            first_name_1 = self.get_cell_value(row_data, 'first_name_1')
            first_name_2 = self.get_cell_value(row_data, 'first_name_2')
            last_name_1 = self.get_cell_value(row_data, 'last_name_1')
            last_name_2 = self.get_cell_value(row_data, 'last_name_2')
            
            # Combine names
            first_name = f"{first_name_1} {first_name_2}".strip() if first_name_1 and first_name_2 else (first_name_1 or first_name_2 or '')
            last_name = f"{last_name_1} {last_name_2}".strip() if last_name_1 and last_name_2 else (last_name_1 or last_name_2 or '')
            
            phone_number1_raw = self.get_cell_value(row_data, 'phone_number1')
            dob_str = self.get_cell_value(row_data, 'dob')
            gender_str = self.get_cell_value(row_data, 'gender')
            role_str = self.get_cell_value(row_data, 'role')
            monthly_salary_str = self.get_cell_value(row_data, 'monthly_salary')
            
            # Validate required fields
            if not all([first_name, last_name, phone_number1_raw, dob_str, gender_str, role_str, monthly_salary_str]):
                raise ValueError("Missing required fields. Check: names, phone_number1, dob, gender, role, monthly_salary")
            
            # Clean and normalize values
            phone_number1 = self.clean_phone_number(phone_number1_raw)
            
            gender = self.normalize_gender(gender_str)
            if not gender:
                raise ValueError(f"Invalid gender value: {gender_str}")
            
            role = self.normalize_role(role_str)
            if not role:
                raise ValueError(f"Invalid role value: {role_str}")
            
            if role not in ['instructor', 'supervisor']:
                raise ValueError(f"Role must be 'instructor' or 'supervisor', got: {role}")
            
            instructor_type = self.normalize_instructor_type(role)
            
            # Parse date
            dob = self.parse_date(dob_str)
            if not dob:
                raise ValueError(f"Invalid date of birth: {dob_str}")
            
            # Parse salary
            monthly_salary = Decimal(str(monthly_salary_str).replace(',', ''))
            
            # Optional fields
            phone_number2 = self.clean_phone_number(self.get_cell_value(row_data, 'phone_number2'))
            if not phone_number2:
                phone_number2 = self.clean_phone_number(self.get_cell_value(row_data, 'phone_number2_alt'))
            
            email = self.get_cell_value(row_data, 'email')
            identity_number = self.get_cell_value(row_data, 'identity_number')
            identity_type = self.normalize_identity_type(self.get_cell_value(row_data, 'identity_type'))
            address = self.get_cell_value(row_data, 'address')
            location = self.get_cell_value(row_data, 'location')
            bio = self.get_cell_value(row_data, 'bio')
            
            # Generate fingerprint ID
            fingerprint_id = self.generate_fingerprint_id()
            
            # Auto-generate secure password (ignore password from Excel)
            password = self.generate_secure_password(length=12)
            
            # Check for duplicates
            duplicate_msg = self.check_duplicate(phone_number1, email, identity_number, fingerprint_id)
            if duplicate_msg:
                print(f"⚠️  Skipped: {duplicate_msg}")
                self.stats['skipped'] += 1
                return
            
            # Use transaction to ensure atomicity
            with transaction.atomic():
                # Create user data
                user_data = {
                    'phone_number1': phone_number1,
                    'first_name': first_name,
                    'last_name': last_name,
                    'dob': dob,
                    'gender': gender,
                    'role': role,
                    'is_verified': False,
                    'is_staff': False,
                }
                
                if phone_number2:
                    user_data['phone_number2'] = phone_number2
                if email:
                    user_data['email'] = email
                if identity_number:
                    user_data['identity_number'] = identity_number
                if identity_type:
                    user_data['identity_type'] = identity_type
                if address:
                    user_data['address'] = address
                if location:
                    user_data['location'] = location
                
                # Create user with password validation enabled
                user = CustomUser.objects.create_user(
                    phone_number1=phone_number1,
                    password=password,
                    **{k: v for k, v in user_data.items() if k != 'phone_number1'}
                )
                
                print(f"✅ Created user: {user.get_full_name()} ({phone_number1})")
                print(f"  🔑 Generated password: {password}")
                
                # Create instructor profile
                instructor_data = {
                    'user': user,
                    'monthly_salary': monthly_salary,
                    'type': instructor_type,
                    'fingerprint_id': fingerprint_id,
                }
                
                if bio:
                    instructor_data['bio'] = bio
                
                # Download and upload ID images to Cloudinary
                nid_front_url = self.get_cell_value(row_data, 'nid_front_url')
                nid_back_url = self.get_cell_value(row_data, 'nid_back_url')
                
                if nid_front_url:
                    print(f"  📄 Processing front ID image...")
                    image_data = self.download_google_drive_image(nid_front_url)
                    if image_data:
                        public_id = self.upload_to_cloudinary(
                            image_data,
                            'instructors/nid',
                            f'{user.id}_nid_front'
                        )
                        if public_id:
                            instructor_data['nid_front'] = public_id
                
                if nid_back_url:
                    print(f"  📄 Processing back ID image...")
                    image_data = self.download_google_drive_image(nid_back_url)
                    if image_data:
                        public_id = self.upload_to_cloudinary(
                            image_data,
                            'instructors/nid',
                            f'{user.id}_nid_back'
                        )
                        if public_id:
                            instructor_data['nid_back'] = public_id
                
                instructor = Instructor.objects.create(**instructor_data)
                print(f"✅ Created instructor profile (Fingerprint: {fingerprint_id})")
                
                # Store imported user data for Excel export
                self.imported_users.append({
                    'full_name': user.get_full_name(),
                    'phone_number1': phone_number1,
                    'phone_number2': phone_number2 or '',
                    'email': email or '',
                    'password': password,
                    'fingerprint_id': fingerprint_id,
                    'role': role,
                    'monthly_salary': str(monthly_salary),
                })
                
                self.stats['success'] += 1
                
        except Exception as e:
            error_msg = f"Row {row_number}: {str(e)}"
            print(f"❌ Error: {error_msg}")
            self.error_messages.append(error_msg)
            self.stats['errors'] += 1
    
    def import_from_excel(self):
        """Main import function"""
        try:
            # Load workbook
            print(f"\nStarting import from: {self.excel_file_path}")
            print("=" * 80)
            
            workbook = openpyxl.load_workbook(self.excel_file_path)
            sheet = workbook.active
            
            # Get header row and map to English field names
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    header = str(cell.value).strip()
                    # Map Arabic header to English field name
                    english_field = self.COLUMN_MAPPING.get(header, header)
                    headers.append(english_field)
            
            print(f"Found {len(headers)} columns")
            print(f"Mapped headers: {', '.join(headers[:5])}...")
            print("=" * 80)
            
            # Process each row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not any(row):
                    continue
                
                self.stats['total'] += 1
                
                # Convert row to dictionary
                row_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = value
                
                print(f"\nProcessing row {row_idx}...")
                self.import_row(row_data, row_idx)
            
            # Print summary
            self.print_summary()
            
        except InvalidFileException:
            print("❌ Error: Invalid Excel file format. Please provide a valid .xlsx or .xls file.")
            sys.exit(1)
        except FileNotFoundError:
            print(f"❌ Error: File not found: {self.excel_file_path}")
            sys.exit(1)
        except Exception as e:
            print(f"❌ Unexpected error: {str(e)}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    def print_summary(self):
        """Print import summary"""
        print("\n" + "=" * 80)
        print("📊 IMPORT SUMMARY")
        print("=" * 80)
        print(f"Total rows processed: {self.stats['total']}")
        print(f"✅ Successfully imported: {self.stats['success']}")
        print(f"⚠️  Skipped (duplicates): {self.stats['skipped']}")
        print(f"❌ Failed (errors): {self.stats['errors']}")
        print("=" * 80)
        
        if self.error_messages:
            print("\n❌ ERROR DETAILS:")
            for msg in self.error_messages:
                print(f"  - {msg}")
    
    def export_passwords_to_excel(self):
        """Export imported users with passwords to Excel file"""
        if not self.imported_users:
            print("\n⚠️  No users were imported, skipping Excel export.")
            return None
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Staff Passwords"
            
            # Define headers (Arabic and English)
            headers = [
                'الاسم الكامل\nFull Name',
                'رقم الهاتف الأساسي\nPrimary Phone',
                'رقم هاتف إضافي\nSecondary Phone',
                'البريد الإلكتروني\nEmail',
                'كلمة المرور\nPassword',
                'معرف البصمة\nFingerprint ID',
                'الدور\nRole',
                'الراتب الشهري\nMonthly Salary',
            ]
            
            # Style for headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Data alignment
            data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            password_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row_idx, user_data in enumerate(self.imported_users, start=2):
                row_data = [
                    user_data['full_name'],
                    user_data['phone_number1'],
                    user_data['phone_number2'],
                    user_data['email'],
                    user_data['password'],
                    user_data['fingerprint_id'],
                    user_data['role'],
                    user_data['monthly_salary'],
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    
                    # Special formatting for password column
                    if col_idx == 5:  # Password column
                        cell.alignment = password_alignment
                        cell.font = Font(bold=True, size=11, color="FF0000")  # Red, bold
                    else:
                        cell.alignment = data_alignment
            
            # Adjust column widths
            column_widths = {
                'A': 30,  # Full Name
                'B': 20,  # Primary Phone
                'C': 20,  # Secondary Phone
                'D': 30,  # Email
                'E': 18,  # Password
                'F': 15,  # Fingerprint ID
                'G': 15,  # Role
                'H': 18,  # Monthly Salary
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Set row height for header
            ws.row_dimensions[1].height = 35
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Add a note/warning at the top
            ws.insert_rows(1)
            warning_cell = ws.cell(row=1, column=1, value="⚠️ تحذير: هذا الملف يحتوي على كلمات مرور حساسة. يرجى حفظه في مكان آمن وعدم مشاركته | WARNING: This file contains sensitive passwords. Keep it secure!")
            warning_cell.font = Font(bold=True, size=11, color="FF0000")
            warning_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            ws.merge_cells('A1:H1')
            warning_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 30
            
            # Save file
            script_dir = Path(__file__).resolve().parent
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'staff_passwords_{timestamp}.xlsx'
            output_file = script_dir / output_filename
            
            wb.save(output_file)
            
            print(f"\n📄 ✅ Password Excel file created successfully!")
            print(f"📁 File location: {output_file}")
            print(f"📊 Contains {len(self.imported_users)} staff members with their passwords")
            print(f"\n🔒 SECURITY REMINDER:")
            print(f"   - This file contains sensitive password information")
            print(f"   - Store it in a secure location")
            print(f"   - Delete it after distributing passwords to staff")
            print(f"   - Do not share via email or unsecured channels")
            
            return output_file
            
        except Exception as e:
            print(f"\n❌ Error creating password Excel file: {str(e)}")
            import traceback
            traceback.print_exc()
            return None


def main():
    """Main entry point"""
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_staff_from_excel.py <excel_file_path>")
        print("\nExample:")
        print("  python scripts/import_staff_from_excel.py staff_data.xlsx")
        print("  python scripts/import_staff_from_excel.py C:/Users/Data/staff.xlsx")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    # Check if file exists
    if not os.path.exists(excel_file):
        # Try relative to script directory
        script_dir = Path(__file__).resolve().parent
        excel_file = script_dir / excel_file
        
        if not os.path.exists(excel_file):
            print(f"❌ Error: File not found: {sys.argv[1]}")
            sys.exit(1)
    
    # Run import
    importer = StaffImporter(excel_file)
    importer.import_from_excel()
    importer.export_passwords_to_excel()


if __name__ == '__main__':
    main()
