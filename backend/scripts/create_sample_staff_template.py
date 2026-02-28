#!/usr/bin/env python
"""
Script to generate a sample Excel template for importing staff data
This creates a template file with example data that users can follow

Usage:
    python scripts/create_sample_staff_template.py
"""

import os
import sys
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def create_sample_template():
    """Create a sample Excel template with example data"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Data"
    
    # Define headers
    headers = [
        'phone_number1',
        'phone_number2',
        'first_name',
        'last_name',
        'email',
        'dob',
        'gender',
        'identity_number',
        'identity_type',
        'address',
        'location',
        'role',
        'monthly_salary',
        'bio',
        'type',
        'fingerprint_id',
        'password'
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Sample data rows
    sample_data = [
        {
            'phone_number1': '+201012345678',
            'phone_number2': '+201087654321',
            'first_name': 'محمد أحمد',
            'last_name': 'علي حسن',
            'email': 'mohamed.ali@example.com',
            'dob': '1990-05-15',
            'gender': 'male',
            'identity_number': '29005151234567',
            'identity_type': 'nid',
            'address': '10 شارع الجامعة، القاهرة',
            'location': 'https://maps.google.com/?q=30.0444,31.2357',
            'role': 'instructor',
            'monthly_salary': 5000.00,
            'bio': 'خبرة 10 سنوات في تدريس الرياضيات',
            'type': 'normal',
            'fingerprint_id': 'FP001',
            'password': 'SecurePass123'
        },
        {
            'phone_number1': '+201098765432',
            'phone_number2': '',
            'first_name': 'أحمد محمود',
            'last_name': 'إبراهيم',
            'email': 'ahmed.ibrahim@example.com',
            'dob': '1985-03-20',
            'gender': 'ذكر',
            'identity_number': '28503201234568',
            'identity_type': 'nid',
            'address': '25 شارع النيل، الجيزة',
            'location': '',
            'role': 'supervisor',
            'monthly_salary': 7000.00,
            'bio': 'مشرف التدريس - خبرة 15 سنة',
            'type': 'supervisor',
            'fingerprint_id': 'FP002',
            'password': ''
        },
        {
            'phone_number1': '+201055555555',
            'phone_number2': '',
            'first_name': 'فاطمة سعيد',
            'last_name': 'محمد',
            'email': 'fatma.mohamed@example.com',
            'dob': '1992-08-10',
            'gender': 'female',
            'identity_number': '29208101234569',
            'identity_type': 'nid',
            'address': '',
            'location': '',
            'role': 'instructor',
            'monthly_salary': 4500.00,
            'bio': '',
            'type': 'normal',
            'fingerprint_id': '',
            'password': ''
        },
        {
            'phone_number1': '+201066666666',
            'phone_number2': '+201099999999',
            'first_name': 'Sarah Ahmed',
            'last_name': 'Mohamed Ali',
            'email': 'sarah.ahmed@example.com',
            'dob': '15/07/1988',
            'gender': 'أنثى',
            'identity_number': '28807151234570',
            'identity_type': 'passport',
            'address': '15 شارع الهرم، الجيزة',
            'location': '',
            'role': 'مدرس',
            'monthly_salary': 5500.00,
            'bio': 'معلمة لغة إنجليزية',
            'type': 'عادي',
            'fingerprint_id': 'FP003',
            'password': 'MyPassword456'
        }
    ]
    
    # Write sample data
    for row_idx, data in enumerate(sample_data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    column_widths = {
        'A': 18,  # phone_number1
        'B': 18,  # phone_number2
        'C': 15,  # first_name
        'D': 15,  # last_name
        'E': 25,  # email
        'F': 12,  # dob
        'G': 10,  # gender
        'H': 18,  # identity_number
        'I': 15,  # identity_type
        'J': 30,  # address
        'K': 40,  # location
        'L': 12,  # role
        'M': 15,  # monthly_salary
        'N': 30,  # bio
        'O': 12,  # type
        'P': 15,  # fingerprint_id
        'Q': 15,  # password
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save file
    script_dir = Path(__file__).resolve().parent
    output_file = script_dir / 'staff_data_template.xlsx'
    
    wb.save(output_file)
    
    print("✅ Sample template created successfully!")
    print(f"📁 File location: {output_file}")
    print("\n📋 Template contains:")
    print(f"   - Header row with all required columns")
    print(f"   - {len(sample_data)} sample data rows")
    print("\n💡 Instructions:")
    print("   1. Open the template file in Excel")
    print("   2. Replace sample data with your actual staff data")
    print("   3. Keep the header row unchanged")
    print("   4. Save the file")
    print("   5. Run: python scripts/import_staff_from_excel.py staff_data_template.xlsx")
    print("\n📖 For detailed column descriptions, see: scripts/IMPORT_STAFF_DATA_GUIDE.md")


if __name__ == '__main__':
    create_sample_template()
